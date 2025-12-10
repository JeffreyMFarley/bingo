from __future__ import annotations

import math
import re
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils.cell import get_column_letter, range_boundaries

try:
    import xlrd  # type: ignore
    from xlrd.xldate import xldate_as_datetime  # type: ignore
except Exception:  # pragma: no cover - optional dependency
    xlrd = None  # type: ignore
    xldate_as_datetime = None  # type: ignore
import random


Value = Any


@dataclass
class SheetTable:
    ref: str
    display_name: Optional[str] = None


@dataclass
class SheetData:
    title: str
    rows: List[List[Any]]
    tables: List[SheetTable]


class ExcelAnalyzer:
    """Analyze Excel workbooks and emit structured metadata."""

    def __init__(
        self,
        db_agent_version: str = "0.1.0",
        sample_values: int = 3,
        enum_threshold: int = 8,
        include_sample_values: bool = True,
        include_enum_values: bool = True,
    ):
        self.db_agent_version = db_agent_version
        self.sample_values = sample_values
        self.enum_threshold = enum_threshold
        self.include_sample_values = include_sample_values
        self.include_enum_values = include_enum_values

    def analyze_workbook(self, workbook_path: str | Path) -> Dict[str, Any]:
        path = Path(workbook_path)
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")

        formula_map = self._scan_for_formulas(path)

        sheet_data_list = self._load_sheet_data(path)
        sheet_reports: List[Dict[str, Any]] = []
        table_profiles: List[Dict[str, Any]] = []

        for idx, sheet in enumerate(sheet_data_list):
            tables = self._extract_tables(sheet, sheet_index=idx)
            has_formulas = formula_map.get(sheet.title, False)
            sheet_class = self._classify_sheet(sheet, tables, has_formulas)
            sheet_reports.append(
                {
                    "sheet_name": sheet.title,
                    "sheet_index": idx,
                    "sheet_type": sheet_class["sheet_type"],
                    "sheet_type_confidence": sheet_class["confidence"],
                    "sheet_type_reason": sheet_class["reason"],
                    "has_formulas": has_formulas,
                    "detected_tables": [self._table_summary(t) for t in tables],
                    "non_tabular_regions": [],
                }
            )
            table_profiles.extend(tables)

        foreign_key_lookup = self._detect_foreign_keys(table_profiles)
        warnings: List[Dict[str, Any]] = []
        table_reports: List[Dict[str, Any]] = []

        for table in table_profiles:
            columns_for_report: List[Dict[str, Any]] = []
            for column in table["columns"]:
                public_column = {
                    k: v for k, v in column.items() if not k.startswith("_")
                }
                columns_for_report.append(public_column)

            primary_key = self._detect_primary_key(table)
            natural_keys = self._detect_natural_keys(table)
            indexes = self._suggest_indexes(table, foreign_key_lookup)
            foreign_keys = foreign_key_lookup.get(table["table_name"], [])

            warnings.extend(table["warnings"])

            table_reports.append(
                {
                    "table_name": table["table_name"],
                    "sheet_name": table["sheet_name"],
                    "range": table["range"],
                    "primary_key": primary_key,
                    "natural_keys": natural_keys,
                    "foreign_keys": foreign_keys,
                    "indexes": indexes,
                    "columns": columns_for_report,
                }
            )

        analysis = {
            "workbook_name": path.name,
            "generated_at": datetime.now(timezone.utc)
            .isoformat()
            .replace("+00:00", "Z"),
            "db_agent_version": self.db_agent_version,
            "sheets": sheet_reports,
            "tables": table_reports,
            "warnings": warnings,
        }
        return analysis

    def _load_sheet_data(self, path: Path) -> List[SheetData]:
        suffix = path.suffix.lower()
        if suffix == ".xls":
            return self._load_xls(path)
        return self._load_xlsx(path)

    def _load_xlsx(self, path: Path) -> List[SheetData]:
        wb = load_workbook(filename=path, data_only=True, read_only=False)
        sheets: List[SheetData] = []
        try:
            for ws in wb.worksheets:
                rows = [list(row) for row in ws.iter_rows(values_only=True)]
                tables = [
                    SheetTable(
                        ref=tbl.ref, display_name=getattr(tbl, "displayName", None)
                    )
                    for tbl in ws.tables.values()
                ]
                sheets.append(SheetData(title=ws.title, rows=rows, tables=tables))
        finally:
            wb.close()
        return sheets

    def _load_xls(self, path: Path) -> List[SheetData]:
        if xlrd is None or xldate_as_datetime is None:
            raise RuntimeError(
                "xlrd is required to read .xls workbooks. Install from requirements.txt."
            )
        book = xlrd.open_workbook(filename=str(path))
        sheets: List[SheetData] = []
        for sheet in book.sheets():
            rows: List[List[Any]] = []
            for row_idx in range(sheet.nrows):
                row_values: List[Any] = []
                for col_idx in range(sheet.ncols):
                    cell = sheet.cell(row_idx, col_idx)
                    value = cell.value
                    if cell.ctype == xlrd.XL_CELL_DATE:
                        try:
                            value = xldate_as_datetime(value, book.datemode)
                        except Exception:
                            pass
                    row_values.append(value)
                rows.append(row_values)
            sheets.append(SheetData(title=sheet.name, rows=rows, tables=[]))
        return sheets

    def _slice_range(
        self, sheet: SheetData, min_row: int, max_row: int, min_col: int, max_col: int
    ) -> List[List[Any]]:
        block: List[List[Any]] = []
        for row_idx in range(min_row, max_row + 1):
            source_row = (
                sheet.rows[row_idx - 1] if 0 <= row_idx - 1 < len(sheet.rows) else []
            )
            block_row: List[Any] = []
            for col_idx in range(min_col, max_col + 1):
                value = (
                    source_row[col_idx - 1]
                    if 0 <= col_idx - 1 < len(source_row)
                    else None
                )
                block_row.append(value)
            block.append(block_row)
        return block

    # ---- sheet classification -----------------------------------------------

    def _sheet_used_bounds(self, sheet: SheetData) -> Optional[Tuple[int, int, int, int]]:
        """Return (min_row, max_row, min_col, max_col) of non-empty cells, or None if empty."""
        has_values = False
        min_row = None
        max_row = None
        min_col = None
        max_col = None

        for r_idx, row in enumerate(sheet.rows, start=1):
            for c_idx, value in enumerate(row, start=1):
                if self._has_cell_value(value):
                    has_values = True
                    if min_row is None or r_idx < min_row:
                        min_row = r_idx
                    if max_row is None or r_idx > max_row:
                        max_row = r_idx
                    if min_col is None or c_idx < min_col:
                        min_col = c_idx
                    if max_col is None or c_idx > max_col:
                        max_col = c_idx
        if not has_values:
            return None
        return min_row, max_row, min_col, max_col

    def _sheet_numeric_fraction(self, sheet: SheetData) -> float:
        non_empty = 0
        numeric_like = 0
        for row in sheet.rows:
            for value in row:
                if self._has_cell_value(value):
                    non_empty += 1
                    if self._is_numeric_like(value):
                        numeric_like += 1
        return (numeric_like / non_empty) if non_empty else 0.0

    def _classify_sheet(
        self,
        sheet: SheetData,
        tables: List[Dict[str, Any]],
        has_formulas: bool,
    ) -> Dict[str, Any]:
        bounds = self._sheet_used_bounds(sheet)
        if bounds is None:
            return {
                "sheet_type": "empty",
                "confidence": 1.0,
                "reason": "No non-empty cells detected.",
            }

        min_row, max_row, min_col, max_col = bounds
        used_rows = max_row - min_row + 1
        used_cols = max_col - min_col + 1
        used_area = used_rows * used_cols

        # table coverage over used area
        total_table_area = 0
        for t in tables:
            t_min_col, t_min_row, t_max_col, t_max_row = range_boundaries(t["range"])
            total_table_area += (t_max_row - t_min_row + 1) * (t_max_col - t_min_col + 1)

        coverage = (total_table_area / used_area) if used_area > 0 else 0.0
        numeric_fraction = self._sheet_numeric_fraction(sheet)

        # --- Heuristics ---

        # 1) No tables at all
        if not tables:
            if has_formulas:
                return {
                    "sheet_type": "spreadsheet",
                    "confidence": 0.85,
                    "reason": "No tabular ranges detected and formulas are present.",
                }

            # Numeric matrix without formulas looks like a pivot/matrix
            if numeric_fraction >= 0.7 and used_rows >= 2 and used_cols >= 2:
                return {
                    "sheet_type": "pivot_table",
                    "confidence": 0.7,
                    "reason": "Single numeric-dense matrix with no formulas detected.",
                }

            return {
                "sheet_type": "spreadsheet",
                "confidence": 0.6,
                "reason": "No tabular ranges detected; sparse or ad-hoc layout.",
            }

        # 2) Multiple tables
        if len(tables) > 1:
            return {
                "sheet_type": "multi_table",
                "confidence": 0.9 if coverage >= 0.6 else 0.75,
                "reason": f"{len(tables)} tabular regions detected on the sheet.",
            }

        # 3) Exactly one table detected
        table = tables[0]
        t_min_col, t_min_row, t_max_col, t_max_row = range_boundaries(table["range"])
        t_rows = t_max_row - t_min_row + 1
        t_cols = t_max_col - t_min_col + 1
        t_area = t_rows * t_cols
        table_coverage = (t_area / used_area) if used_area > 0 else 0.0

        # Pivot-like: dense numeric matrix in single table
        if numeric_fraction >= 0.8 and t_rows >= 3 and t_cols >= 3 and not has_formulas:
            return {
                "sheet_type": "pivot_table",
                "confidence": 0.75,
                "reason": "Single dense numeric table with matrix-like structure.",
            }

        # Full-table: single table dominates the used area
        if table_coverage >= 0.7:
            return {
                "sheet_type": "full_table",
                "confidence": 0.9 if not has_formulas else 0.8,
                "reason": "Single detected table covers most of the used cells.",
            }

        # Otherwise: table plus other spreadsheet content
        if has_formulas:
            return {
                "sheet_type": "spreadsheet",
                "confidence": 0.7,
                "reason": "Single table plus significant formula-based content.",
            }

        return {
            "sheet_type": "full_table",
            "confidence": 0.6,
            "reason": "Single table detected but it does not completely cover the used area.",
        }

    # ----- table extraction -------------------------------------------------

    def _extract_tables(
        self, sheet: SheetData, sheet_index: int
    ) -> List[Dict[str, Any]]:
        tables: List[Dict[str, Any]] = []
        table_counter = 1
        existing_ranges: List[str] = []

        if sheet.tables:
            for defined_table in sheet.tables:
                table = self._build_table_from_range(
                    sheet=sheet,
                    sheet_index=sheet_index,
                    ref=defined_table.ref,
                    table_name=defined_table.display_name
                    or f"{sheet.title}_Table{table_counter}",
                    confidence=0.98,
                    notes=(
                        [f"Excel table '{defined_table.display_name}' detected"]
                        if defined_table.display_name
                        else ["Excel table detected"]
                    ),
                )
                if table:
                    tables.append(table)
                    existing_ranges.append(table["range"])
                    table_counter += 1

        inferred_ranges = self._infer_table_ranges(sheet)
        for ref in inferred_ranges:
            if any(self._ranges_overlap(ref, existing) for existing in existing_ranges):
                continue
            table = self._build_table_from_range(
                sheet=sheet,
                sheet_index=sheet_index,
                ref=ref,
                table_name=f"{sheet.title}_Table{table_counter}",
                confidence=0.8,
                notes=[f"Contiguous data block detected at {ref}"],
            )
            if table:
                tables.append(table)
                existing_ranges.append(table["range"])
                table_counter += 1

        return tables

    def _build_table_from_range(
        self,
        sheet: SheetData,
        sheet_index: int,
        ref: str,
        table_name: str,
        confidence: float,
        notes: Optional[List[str]] = None,
    ) -> Optional[Dict[str, Any]]:
        min_col, min_row, max_col, max_row = range_boundaries(ref)
        total_rows = max_row - min_row + 1
        if total_rows < 2:
            return None

        values = self._slice_range(sheet, min_row, max_row, min_col, max_col)
        header = values[0]
        data_rows = values[1:]

        if not any(self._has_cell_value(v) for v in header):
            return None

        normalized_headers = self._normalize_header_row(header)
        row_count = len(data_rows)
        column_count = len(normalized_headers)
        data_start_row = min_row + 1

        table = {
            "table_name": table_name,
            "sheet_name": sheet.title,
            "sheet_index": sheet_index,
            "range": ref,
            "header_row": min_row,
            "data_start_row": data_start_row,
            "row_count": row_count,
            "column_count": column_count,
            "confidence": round(confidence, 2),
            "notes": (notes or []) + [f"Header row detected at row {min_row}"],
            "columns": [],
            "warnings": [],
        }

        for idx, header_value in enumerate(normalized_headers):
            column_letter = get_column_letter(min_col + idx)
            col_values = [row[idx] if idx < len(row) else None for row in data_rows]
            column_profile = self._profile_column(
                column_name=header_value,
                excel_column=column_letter,
                values=col_values,
                data_start_row=data_start_row,
                row_count=row_count,
                sheet_name=sheet.title,
                table_name=table_name,
            )
            table["columns"].append(column_profile)
            table["warnings"].extend(
                self._column_warnings(table_name, header_value, column_profile)
            )

        return table

    def _infer_table_ranges(self, sheet: SheetData) -> List[str]:
        rows = sheet.rows
        if not rows:
            return []

        blocks: List[Dict[str, int]] = []
        current_block: Optional[Dict[str, int]] = None

        for idx, row in enumerate(rows, start=1):
            columns_with_values = [
                i for i, value in enumerate(row, start=1) if self._has_cell_value(value)
            ]
            if columns_with_values:
                min_col = min(columns_with_values)
                max_col = max(columns_with_values)
                if current_block is None:
                    current_block = {
                        "start": idx,
                        "end": idx,
                        "min_col": min_col,
                        "max_col": max_col,
                    }
                else:
                    current_block["end"] = idx
                    current_block["min_col"] = min(current_block["min_col"], min_col)
                    current_block["max_col"] = max(current_block["max_col"], max_col)
            else:
                if current_block:
                    blocks.append(current_block)
                    current_block = None

        if current_block:
            blocks.append(current_block)

        ranges: List[str] = []
        for block in blocks:
            if block["end"] - block["start"] < 1:
                continue
            start_col_letter = get_column_letter(block["min_col"])
            end_col_letter = get_column_letter(block["max_col"])
            ranges.append(
                f"{start_col_letter}{block['start']}:{end_col_letter}{block['end']}"
            )

        return ranges

    # ----- formula detection ----------------------------------------------------

    def _scan_for_formulas(self, path: Path) -> Dict[str, bool]:
        """Return {sheet_title: has_formulas} without loading values."""
        suffix = path.suffix.lower()
        if suffix == ".xls":
            return self._scan_for_formulas_xls(path)
        return self._scan_for_formulas_xlsx(path)

    def _scan_for_formulas_xlsx(self, path: Path) -> Dict[str, bool]:
        result: Dict[str, bool] = {}
        wb = load_workbook(filename=path, data_only=False, read_only=True)
        try:
            for ws in wb.worksheets:
                has_formulas = False
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if isinstance(v, str) and v.startswith("="):
                            has_formulas = True
                            break
                    if has_formulas:
                        break
                result[ws.title] = has_formulas
        finally:
            wb.close()
        return result

    def _scan_for_formulas_xls(self, path: Path) -> Dict[str, bool]:
        """
        Excel .xls files read through xlrd do not expose formula information.
        xlrd returns only the last cached value, and no cell type indicates
        whether it was originally a formula.

        Therefore, we conservatively mark all sheets as having no formulas.
        """
        if xlrd is None:
            return {}

        book = xlrd.open_workbook(filename=str(path), formatting_info=False)
        result = {sheet.name: False for sheet in book.sheets()}
        return result


    # ----- column analysis --------------------------------------------------

    def _profile_column(
        self,
        column_name: str,
        excel_column: str,
        values: Sequence[Value],
        data_start_row: int,
        row_count: int,
        sheet_name: str,
        table_name: str,
    ) -> Dict[str, Any]:
        non_null_values = [v for v in values if self._has_cell_value(v)]
        normalized_values = [
            self._normalize_value(v)
            for v in non_null_values
            if self._normalize_value(v) is not None
        ]
        unique_values = set(normalized_values)
        nullable = len(non_null_values) != len(values)
        unique_ratio = (
            (len(unique_values) / len(non_null_values)) if non_null_values else 0.0
        )

        data_type = self._infer_data_type(column_name, non_null_values, unique_values)
        inferred_constraints = self._infer_constraints(
            data_type, non_null_values, unique_values, column_name
        )

        if self.include_sample_values and self.sample_values > 0:
            sample_values_raw = random.sample(
                non_null_values, min(self.sample_values, len(non_null_values))
            )
            sample_values = [self._stringify_value(v) for v in sample_values_raw]
        else:
            sample_values = []
        source_cells = (
            f"{excel_column}{data_start_row}:{excel_column}{data_start_row + row_count - 1}"
            if row_count > 0
            else f"{excel_column}{data_start_row}"
        )

        profile = {
            "column_name": column_name,
            "excel_column": excel_column,
            "data_type": data_type,
            "nullable": nullable,
            "inferred_constraints": inferred_constraints,
            "sample_values": sample_values,
            "description": f"Column '{column_name}' from table '{table_name}' on sheet '{sheet_name}'.",
            "source_cells": source_cells,
            "_stats": {
                "non_null_count": len(non_null_values),
                "total_count": len(values),
                "unique_values": unique_values,
                "unique_ratio": unique_ratio,
                "enum_candidate": data_type == "enum"
                or len(unique_values)
                <= min(self.enum_threshold, max(1, len(non_null_values))),
                "is_numeric": data_type in {"integer", "decimal", "currency"},
                "is_text": data_type in {"string", "enum"},
                "is_date": data_type == "date",
                "nullable": nullable,
                "normalized_values": unique_values,
            },
        }
        return profile

    def _infer_data_type(
        self, column_name: str, values: Sequence[Value], unique_values: Iterable[str]
    ) -> str:
        if not values:
            return "string"

        header = column_name.lower()
        if all(isinstance(v, bool) for v in values):
            return "boolean"

        if all(self._is_date_like(v) for v in values):
            return "date"

        if all(self._is_numeric_like(v) for v in values):
            if all(self._is_integer_like(v) for v in values):
                return "integer"
            if any(
                keyword in header
                for keyword in ("amount", "total", "subtotal", "price", "cost", "tax")
            ):
                return "currency"
            return "decimal"

        if self._looks_like_enum(header, values, unique_values):
            return "enum"

        return "string"

    def _infer_constraints(
        self,
        data_type: str,
        values: Sequence[Value],
        unique_values: Iterable[str],
        column_name: str,
    ) -> Dict[str, Any]:
        constraints: Dict[str, Any] = {}
        unique = len(values) > 0 and len(set(unique_values)) == len(values)

        if data_type in {"integer", "decimal", "currency"}:
            numeric_values = [
                self._to_number(v) for v in values if self._to_number(v) is not None
            ]
            if numeric_values:
                constraints["min"] = min(numeric_values)
                constraints["max"] = max(numeric_values)
        elif data_type == "date":
            date_values = [
                self._to_date(v) for v in values if self._to_date(v) is not None
            ]
            if date_values:
                constraints["min"] = min(date_values).isoformat()
                constraints["max"] = max(date_values).isoformat()
        elif data_type == "string":
            lengths = [len(str(v)) for v in values]
            if lengths:
                constraints["max_length"] = max(lengths)
            if any(self._looks_like_email(str(v)) for v in values):
                constraints["pattern"] = "email"
        elif data_type == "enum":
            if self.include_enum_values:
                samples = sorted({self._stringify_value(v) for v in values})
                constraints["allowed_values"] = samples[:15]

        if unique:
            constraints["unique"] = True

        return constraints

    def _column_warnings(
        self, table_name: str, column_name: str, column_profile: Dict[str, Any]
    ) -> List[Dict[str, Any]]:
        stats = column_profile["_stats"]
        warnings: List[Dict[str, Any]] = []
        if stats["enum_candidate"] and stats["non_null_count"] > 0:
            warnings.append(
                {
                    "code": "ENUM_GUESS",
                    "severity": "info",
                    "message": f"Column '{column_name}' appears to contain a small set of values.",
                    "table": table_name,
                    "column": column_name,
                    "examples": column_profile["sample_values"],
                }
            )
        return warnings

    # ----- key detection ----------------------------------------------------

    def _detect_primary_key(self, table: Dict[str, Any]) -> Dict[str, Any]:
        best_candidate: Optional[Tuple[Dict[str, Any], float]] = None
        for column in table["columns"]:
            stats = column["_stats"]
            if stats["non_null_count"] == 0:
                continue
            if stats["unique_ratio"] < 0.95 or stats["nullable"]:
                continue

            confidence = stats["unique_ratio"]
            if "id" in column["column_name"].lower():
                confidence += 0.05
            if stats["is_numeric"]:
                confidence += 0.02

            confidence = min(confidence, 0.99)
            if best_candidate is None or confidence > best_candidate[1]:
                best_candidate = (column, confidence)

        if best_candidate:
            column, confidence = best_candidate
            return {
                "columns": [column["column_name"]],
                "confidence": round(confidence, 2),
                "strategy": "unique_values",
            }

        return {"columns": [], "confidence": 0.0, "strategy": "not_detected"}

    def _detect_natural_keys(self, table: Dict[str, Any]) -> List[Dict[str, Any]]:
        natural_keys: List[Dict[str, Any]] = []
        for column in table["columns"]:
            stats = column["_stats"]
            if stats["non_null_count"] == 0:
                continue
            if stats["unique_ratio"] >= 0.85 and column["data_type"] in {
                "string",
                "enum",
            }:
                natural_keys.append(
                    {
                        "columns": [column["column_name"]],
                        "confidence": round(stats["unique_ratio"], 2),
                        "strategy": "unique_values",
                    }
                )
        return natural_keys

    def _detect_foreign_keys(
        self, tables: Sequence[Dict[str, Any]]
    ) -> Dict[str, List[Dict[str, Any]]]:
        fk_map: Dict[str, List[Dict[str, Any]]] = {
            table["table_name"]: [] for table in tables
        }

        for table in tables:
            for column in table["columns"]:
                stats = column["_stats"]
                if stats["non_null_count"] == 0:
                    continue
                normalized_values = stats["normalized_values"]
                if not normalized_values:
                    continue

                for target_table in tables:
                    if target_table is table:
                        continue
                    for target_column in target_table["columns"]:
                        target_stats = target_column["_stats"]
                        if target_stats["non_null_count"] == 0:
                            continue
                        overlap = normalized_values.intersection(
                            target_stats["normalized_values"]
                        )
                        if not overlap:
                            continue
                        overlap_ratio = len(overlap) / max(1, len(normalized_values))
                        name_score = self._foreign_key_name_score(
                            column["column_name"],
                            target_column["column_name"],
                            target_table["table_name"],
                        )
                        if overlap_ratio < 0.5 and name_score < 0.6:
                            continue
                        confidence = min(0.99, 0.5 * overlap_ratio + 0.5 * name_score)
                        fk_map[table["table_name"]].append(
                            {
                                "columns": [column["column_name"]],
                                "references_table": target_table["table_name"],
                                "references_columns": [target_column["column_name"]],
                                "confidence": round(confidence, 2),
                                "strategy": "name_similarity_and_value_overlap",
                            }
                        )
        return fk_map

    def _foreign_key_name_score(
        self, column_name: str, target_column_name: str, target_table_name: str
    ) -> float:
        cname = column_name.lower()
        target = target_column_name.lower()
        table = target_table_name.lower()

        if cname == target:
            return 0.9
        if cname.endswith("id"):
            prefix = cname[:-2]
            if prefix == table or prefix == target.rstrip("id"):
                return 0.85
        if cname.startswith(table):
            return 0.7
        if cname == f"{table}_id":
            return 0.9
        if cname == f"{target}_id":
            return 0.85
        return 0.4 if target in cname or table in cname else 0.2

    def _suggest_indexes(
        self, table: Dict[str, Any], fk_map: Dict[str, List[Dict[str, Any]]]
    ) -> List[Dict[str, Any]]:
        suggestions: List[Dict[str, Any]] = []
        fk_columns = {
            tuple(fk["columns"]) for fk in fk_map.get(table["table_name"], [])
        }

        for fk in fk_map.get(table["table_name"], []):
            suggestions.append(
                {
                    "columns": fk["columns"],
                    "reason": "foreign key candidate",
                    "confidence": fk["confidence"],
                }
            )

        date_columns = [
            col["column_name"] for col in table["columns"] if col["data_type"] == "date"
        ]
        for column_name in date_columns:
            suggestions.append(
                {
                    "columns": [column_name],
                    "reason": "date column useful for sorting/filtering",
                    "confidence": 0.7,
                }
            )

        name_columns = [
            col["column_name"]
            for col in table["columns"]
            if "name" in col["column_name"].lower()
        ]
        if len(name_columns) >= 2:
            pair = name_columns[:2]
            if tuple(pair) not in fk_columns:
                suggestions.append(
                    {
                        "columns": pair,
                        "reason": "text columns likely involved in searches",
                        "confidence": 0.65,
                    }
                )

        return suggestions

    # ----- helpers ----------------------------------------------------------

    def _table_summary(self, table: Dict[str, Any]) -> Dict[str, Any]:
        return {
            "table_name": table["table_name"],
            "range": table["range"],
            "header_row": table["header_row"],
            "data_start_row": table["data_start_row"],
            "row_count": table["row_count"],
            "column_count": table["column_count"],
            "confidence": table["confidence"],
            "notes": table["notes"],
        }

    def _ranges_overlap(self, ref_a: str, ref_b: str) -> bool:
        min_col_a, min_row_a, max_col_a, max_row_a = range_boundaries(ref_a)
        min_col_b, min_row_b, max_col_b, max_row_b = range_boundaries(ref_b)
        rows_overlap = not (max_row_a < min_row_b or min_row_a > max_row_b)
        cols_overlap = not (max_col_a < min_col_b or min_col_a > max_col_b)
        return rows_overlap and cols_overlap

    def _normalize_header_row(self, header: Sequence[Value]) -> List[str]:
        normalized: List[str] = []
        seen: Dict[str, int] = {}
        for idx, value in enumerate(header, start=1):
            text = str(value).strip() if self._has_cell_value(value) else ""
            if not text:
                text = f"Column{idx}"
            key = text.lower()
            if key in seen:
                seen[key] += 1
                text = f"{text}_{seen[key]}"
            else:
                seen[key] = 1
            normalized.append(text)
        return normalized

    def _has_cell_value(self, value: Any) -> bool:
        if value is None:
            return False
        if isinstance(value, str):
            return value.strip() != ""
        return True

    def _normalize_value(self, value: Value) -> Optional[str]:
        if not self._has_cell_value(value):
            return None
        if isinstance(value, bool):
            return str(value).lower()
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, float):
            if math.isfinite(value) and value.is_integer():
                return str(int(value))
            return f"{value}"
        return str(value).strip()

    def _stringify_value(self, value: Value) -> str:
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return "" if value is None else str(value)

    def _is_numeric_like(self, value: Value) -> bool:
        if isinstance(value, (int, float)) and math.isfinite(value):
            return True
        if isinstance(value, str):
            stripped = value.replace(",", "").strip()
            if not stripped:
                return False
            try:
                float(stripped)
                return True
            except ValueError:
                return False
        return False

    def _is_integer_like(self, value: Value) -> bool:
        if isinstance(value, int):
            return True
        if isinstance(value, float) and value.is_integer():
            return True
        if isinstance(value, str):
            stripped = value.strip().replace(",", "")
            return stripped.isdigit() or (
                stripped.startswith("-") and stripped[1:].isdigit()
            )
        return False

    def _looks_like_enum(
        self, header: str, values: Sequence[Value], unique_values: Iterable[str]
    ) -> bool:
        if any(
            keyword in header for keyword in ("status", "type", "category", "state")
        ):
            return True
        unique_count = len(set(unique_values))
        return unique_count <= min(self.enum_threshold, max(1, len(values)))

    def _looks_like_email(self, value: str) -> bool:
        return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", value))

    def _is_date_like(self, value: Value) -> bool:
        if isinstance(value, (datetime, date)):
            return True
        if isinstance(value, str):
            try:
                datetime.fromisoformat(value)
                return True
            except ValueError:
                return bool(re.match(r"^\d{4}[-/]\d{2}[-/]\d{2}$", value.strip()))
        return False

    def _to_number(self, value: Value) -> Optional[float]:
        if isinstance(value, (int, float)) and math.isfinite(value):
            return float(value)
        if isinstance(value, str):
            stripped = value.replace(",", "").strip()
            if not stripped:
                return None
            try:
                return float(stripped)
            except ValueError:
                return None
        return None

    def _to_date(self, value: Value) -> Optional[date]:
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            try:
                parsed = datetime.fromisoformat(value)
                return parsed.date()
            except ValueError:
                match = re.match(r"^(\d{4})[-/](\d{2})[-/](\d{2})$", value.strip())
                if match:
                    year, month, day = map(int, match.groups())
                    return date(year, month, day)
        return None
